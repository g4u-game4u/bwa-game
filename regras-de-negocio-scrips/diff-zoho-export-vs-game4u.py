#!/usr/bin/env python3
"""
Cruza export Zoho (xlsx em historico-zoho-export/) com GET /user-action/search no Game4U.

Para cada linha do export, calcula o integration_id estável (mesma regra do run-api-scripts.mjs)
e indica se existe user_action no Game4U e qual o status.

Uso (na raiz do repo, com .env):
  python regras-de-negocio-scrips/diff-zoho-export-vs-game4u.py

Variáveis: G4U_API_BASE, client_id ou CLIENT_ID, G4U_ACCESS_TOKEN ou G4U_ADMIN_EMAIL/PASSWORD.
Opcional: G4U_DIFF_CREATED_START, G4U_DIFF_CREATED_END (ISO), G4U_DIFF_PER_DEAL=1 (força search por delivery_id),
--out caminho.csv
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from collections import Counter
import unicodedata
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.parse import urlencode
from urllib.request import Request, urlopen

try:
    from dotenv import load_dotenv
except ImportError:
    load_dotenv = None

try:
    import openpyxl
except ImportError:
    print("Instale openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_EXPORT_DIR = REPO_ROOT / "historico-zoho-export"
MAP_PATH = REPO_ROOT / "src" / "app" / "config" / "zoho-crm-action-map.json"


def norm_key(s: str) -> str:
    s = unicodedata.normalize("NFKD", str(s or ""))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return s.lower().strip()


def find_col_index(headers: list[Any], *candidates: str) -> int | None:
    """Índice da coluna cujo header normalizado contém algum candidato."""
    normed = [(i, norm_key(h)) for i, h in enumerate(headers)]
    for cand in candidates:
        c = norm_key(cand)
        for i, hn in normed:
            if c in hn or hn in c:
                return i
    return None


def find_col_index_exact_norm(headers: list[Any], exact_label: str) -> int | None:
    """Match exato do header normalizado (evita 'Nome Negócio' casar com 'Nome Negócio.id')."""
    want = norm_key(exact_label)
    for i, h in enumerate(headers):
        if norm_key(h) == want:
            return i
    return None


def resolve_stage_to_template(
    stage: str, stage_map: dict[str, str], norm_map: dict[str, tuple[str, str]]
) -> tuple[str | None, str | None, str]:
    """
    Retorna (chave_map_usada, template_id, nota).
    norm_map: norm(stage_title) -> (chave_original, template_id)
    """
    if not stage or not str(stage).strip():
        return None, None, "SEM_STAGE"
    s = str(stage).strip()
    if s in stage_map:
        return s, stage_map[s], ""
    nk = norm_key(s)
    if nk in norm_map:
        orig, tid = norm_map[nk]
        return orig, tid, f"NORMALIZED_STAGE:{orig}"
    return None, None, f"UNMAPPED_STAGE:{s}"


def zcrm_to_id(raw: Any) -> str | None:
    if raw is None:
        return None
    s = str(raw).strip()
    if not s or s.lower() == "none":
        return None
    m = re.search(r"(\d{10,})", s)
    return m.group(1) if m else None


def stable_stage_integration_id(deal_id: str, action_template_id: str) -> str:
    return f"zoho-deal-{deal_id}-action-{action_template_id}"


def stable_jur_integration_id(deal_id: str, action_template_id: str) -> str:
    return f"zoho-deal-{deal_id}-jur-action-{action_template_id}"


def zoho_activity_matches_tag_flow_key(subject: str, export_tag: str | None, tag_flow_key: str) -> bool:
    """Espelha zohoActivityMatchesTagFlowKey (run-api-scripts.mjs) + coluna Tag do export."""
    if not tag_flow_key:
        return False
    blob = json.dumps(
        {"Subject": subject or "", "Tag": export_tag or ""}, ensure_ascii=False
    )
    if tag_flow_key in blob:
        return True
    subj = (subject or "").strip()
    subjl = subj.lower()
    tag_m = re.search(r"Tag:\s*(.+)$", tag_flow_key, re.I)
    if tag_m:
        label = tag_m.group(1).strip()
        et = (export_tag or "").strip()
        if et and et.lower() == label.lower():
            return True
        if len(label) >= 2 and label.lower() in subjl:
            return True
    before_tag = re.split(r"\s*-\s*Tag:", tag_flow_key, maxsplit=1, flags=re.I)[0]
    before_tag = re.sub(r"^\*\s*", "", before_tag).strip()
    if len(before_tag) >= 5 and before_tag.lower() in subjl:
        return True
    return False


def resolve_jur_action(subject: str, export_tag: str | None, tag_map: dict[str, str]) -> tuple[str | None, str | None]:
    entries = sorted(tag_map.items(), key=lambda x: len(x[0]), reverse=True)
    for k, action_id in entries:
        if zoho_activity_matches_tag_flow_key(subject, export_tag, k):
            return k, action_id
    return None, None


def load_action_map(path: Path) -> dict[str, Any]:
    with path.open(encoding="utf-8") as f:
        return json.load(f)


def load_env(repo: Path) -> None:
    env_file = repo / ".env"
    if load_dotenv and env_file.is_file():
        load_dotenv(env_file)


def env_get(name: str, default: str | None = None) -> str | None:
    import os

    v = os.environ.get(name)
    if v is not None and str(v).strip():
        return str(v).strip()
    return default


def game4u_login(base: str, client_id: str, email: str, password: str) -> str:
    url = base.rstrip("/") + "/auth/login"
    body = json.dumps({"email": email, "password": password}).encode("utf-8")
    req = Request(
        url,
        data=body,
        method="POST",
        headers={
            "client_id": client_id,
            "Content-Type": "application/json",
            "Accept": "application/json",
        },
    )
    with urlopen(req, timeout=120) as resp:
        data = json.loads(resp.read().decode("utf-8"))
    token = data.get("access_token") or (data.get("data") or {}).get("access_token")
    if not token:
        token = (data.get("token") or {}).get("access_token")
    if not token:
        raise RuntimeError("Login sem access_token reconhecido")
    return str(token)


def game4u_get(base: str, path_with_qs: str, client_id: str, token: str) -> dict[str, Any]:
    url = base.rstrip("/") + (path_with_qs if path_with_qs.startswith("/") else "/" + path_with_qs)
    req = Request(
        url,
        method="GET",
        headers={
            "client_id": client_id,
            "Authorization": f"Bearer {token}",
            "Accept": "application/json",
        },
    )
    with urlopen(req, timeout=120) as resp:
        return json.loads(resp.read().decode("utf-8"))


def unwrap_items(j: dict[str, Any]) -> list[dict[str, Any]]:
    if isinstance(j.get("items"), list):
        return j["items"]
    if isinstance(j.get("data"), list):
        return j["data"]
    return []


def fetch_all_user_actions(
    base: str,
    client_id: str,
    token: str,
    created_start: str,
    created_end: str,
    statuses: list[str],
    limit: int,
    max_pages: int,
) -> list[dict[str, Any]]:
    """GET /user-action/search sem delivery_id (índice global por integration_id)."""
    all_rows: list[dict[str, Any]] = []
    for status in statuses:
        page = 1
        page_token: str | None = None
        for _ in range(max_pages):
            q: dict[str, str] = {
                "created_at_start": created_start,
                "created_at_end": created_end,
                "dismissed": "false",
                "limit": str(min(max(limit, 1), 500)),
                "status": status,
            }
            if page_token:
                q["page_token"] = page_token
            else:
                q["page"] = str(page)
            qs = "?" + urlencode(q)
            path = "/user-action/search" + qs
            try:
                j = game4u_get(base, path, client_id, token)
            except HTTPError as e:
                body = e.read().decode("utf-8", errors="replace")[:2000]
                raise RuntimeError(f"HTTP {e.code} em {path}: {body}") from e
            chunk = unwrap_items(j)
            all_rows.extend(chunk)
            if not chunk:
                break
            next_tok = j.get("next_page_token") or j.get("nextPageToken")
            if next_tok:
                page_token = str(next_tok)
                continue
            page_token = None
            total_pages = j.get("total_pages") or j.get("totalPages")
            cur = j.get("page")
            if isinstance(total_pages, int) and isinstance(cur, int) and cur < total_pages:
                page = cur + 1
                continue
            if len(chunk) >= limit:
                page += 1
                continue
            break
    return all_rows


def fetch_user_actions_per_deal(
    base: str,
    client_id: str,
    token: str,
    deal_ids: list[str],
    created_start: str,
    created_end: str,
    statuses: list[str],
    limit: int,
    max_pages: int,
) -> list[dict[str, Any]]:
    """GET /user-action/search com delivery_id para cada deal (quando a API exige filtro)."""
    all_rows: list[dict[str, Any]] = []
    for did in deal_ids:
        for st in statuses:
            page = 1
            page_token: str | None = None
            for _ in range(max_pages):
                q: dict[str, str] = {
                    "created_at_start": created_start,
                    "created_at_end": created_end,
                    "dismissed": "false",
                    "limit": str(min(max(limit, 1), 500)),
                    "status": st,
                    "delivery_id": str(did),
                }
                if page_token:
                    q["page_token"] = page_token
                else:
                    q["page"] = str(page)
                qs = "?" + urlencode(q)
                path = "/user-action/search" + qs
                try:
                    j = game4u_get(base, path, client_id, token)
                except HTTPError:
                    break
                chunk = unwrap_items(j)
                all_rows.extend(chunk)
                if not chunk:
                    break
                next_tok = j.get("next_page_token") or j.get("nextPageToken")
                if next_tok:
                    page_token = str(next_tok)
                    continue
                page_token = None
                total_pages = j.get("total_pages") or j.get("totalPages")
                cur = j.get("page")
                if isinstance(total_pages, int) and isinstance(cur, int) and cur < total_pages:
                    page = cur + 1
                    continue
                if len(chunk) >= limit:
                    page += 1
                    continue
                break
    return all_rows


def index_by_integration_id(rows: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    """Uma linha por integration_id; prioriza DONE > PENDING > CANCELLED (mesmo id em vários status)."""
    prio = {"CANCELLED": 0, "PENDING": 1, "DONE": 2}

    def rank(r: dict[str, Any]) -> int:
        return prio.get(str(r.get("status") or "").strip().upper(), 0)

    out: dict[str, dict[str, Any]] = {}
    for r in rows:
        iid = r.get("integration_id") or r.get("integrationId")
        if not iid:
            continue
        s = str(iid).strip()
        prev = out.get(s)
        if not prev or rank(r) >= rank(prev):
            out[s] = r
    return out


def pick_xlsx_files(export_dir: Path) -> list[tuple[str, Path]]:
    files: list[tuple[str, Path]] = []
    for p in sorted(export_dir.glob("*.xlsx")):
        name = p.name.lower()
        if "cs mar" in name or ("cs" in name and "mar" in name and "cobran" not in name):
            files.append(("CS", p))
        elif "cobran" in name and "mar" in name:
            files.append(("Cobrança", p))
        elif "jur" in name and "mar" in name:
            files.append(("Jurídico", p))
    return files


def read_stage_sheet(
    kind: str,
    path: Path,
    stage_map: dict[str, str],
    norm_stage_map: dict[str, tuple[str, str]],
) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    headers = list(next(rows_iter))
    idx_id = find_col_index(headers, "id do registro", "record id")
    idx_deal = find_col_index_exact_norm(headers, "Nome Negócio.id") or find_col_index(
        headers, "nome negocio.id"
    )
    idx_deal_name = find_col_index_exact_norm(headers, "Nome Negócio")
    idx_stage = find_col_index(headers, "stage")
    idx_mod = find_col_index(headers, "hora da modifica", "modificacao")
    idx_fluxo = find_col_index(headers, "fluxo")
    out: list[dict[str, Any]] = []
    row_num = 1
    for tup in rows_iter:
        row_num += 1
        if not tup or all(v is None or str(v).strip() == "" for v in tup):
            continue
        def g(i: int | None):
            if i is None or i >= len(tup):
                return None
            return tup[i]

        deal_id = zcrm_to_id(g(idx_deal))
        stage_raw = g(idx_stage)
        stage = (str(stage_raw).strip() if stage_raw is not None else "") or ""
        _map_key, tpl, map_note = resolve_stage_to_template(stage, stage_map, norm_stage_map)
        int_id = stable_stage_integration_id(deal_id, tpl) if (deal_id and tpl) else None
        out.append(
            {
                "source_kind": kind,
                "source_file": path.name,
                "row": row_num,
                "export_record_id": g(idx_id),
                "deal_id": deal_id,
                "deal_name": g(idx_deal_name),
                "stage": stage,
                "modified": g(idx_mod),
                "fluxo": g(idx_fluxo),
                "expected_template": tpl,
                "expected_integration_id": int_id,
                "map_note": map_note,
            }
        )
    wb.close()
    return out


def read_jur_sheet(path: Path, tag_map: dict[str, str]) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    headers = list(next(rows_iter))
    idx_id = find_col_index(headers, "id do registro")
    idx_deal = find_col_index_exact_norm(headers, "Relacionado A.id") or find_col_index(
        headers, "relacionado a.id"
    )
    idx_deal_name = find_col_index_exact_norm(headers, "Relacionado A")
    idx_subject = find_col_index(headers, "assunto")
    idx_status = find_col_index(headers, "status")
    idx_tag = find_col_index(headers, "tag")
    idx_closed = find_col_index(headers, "fechamento", "horario de fechamento")
    out: list[dict[str, Any]] = []
    row_num = 1
    for tup in rows_iter:
        row_num += 1
        if not tup or all(v is None or str(v).strip() == "" for v in tup):
            continue

        def g(i: int | None):
            if i is None or i >= len(tup):
                return None
            return tup[i]

        deal_id = zcrm_to_id(g(idx_deal))
        subject = str(g(idx_subject) or "").strip()
        tag = g(idx_tag)
        tag_s = str(tag).strip() if tag is not None else ""
        map_key, tpl = resolve_jur_action(subject, tag_s or None, tag_map)
        int_id = stable_jur_integration_id(deal_id, tpl) if (deal_id and tpl) else None
        note = ""
        if not tpl:
            note = f"UNMAPPED_JUR_TAG:{tag_s or '-'}; subject_len={len(subject)}"
        out.append(
            {
                "source_kind": "Jurídico",
                "source_file": path.name,
                "row": row_num,
                "export_record_id": g(idx_id),
                "deal_id": deal_id,
                "deal_name": g(idx_deal_name),
                "stage": subject,
                "zoho_tag": tag_s,
                "zoho_task_status": g(idx_status),
                "modified": g(idx_closed),
                "fluxo": "",
                "expected_template": tpl,
                "expected_integration_id": int_id,
                "map_note": note,
                "map_key": map_key or "",
            }
        )
    wb.close()
    return out


def main() -> int:
    parser = argparse.ArgumentParser(description="Diff export Zoho xlsx vs Game4U user_action")
    parser.add_argument(
        "--export-dir",
        type=Path,
        default=DEFAULT_EXPORT_DIR,
        help="Pasta com os xlsx de histórico",
    )
    parser.add_argument(
        "--out",
        type=Path,
        default=None,
        help="CSV de saída (default: historico-zoho-export/diff-zoho-game4u-TIMESTAMP.csv)",
    )
    parser.add_argument(
        "--created-start",
        default=None,
        help="ISO created_at_start para /user-action/search",
    )
    parser.add_argument(
        "--created-end",
        default=None,
        help="ISO created_at_end para /user-action/search",
    )
    parser.add_argument(
        "--dry-parse",
        action="store_true",
        help="Só lê xlsx e mapa; não chama API",
    )
    args = parser.parse_args()

    load_env(REPO_ROOT)
    created_start = args.created_start or env_get(
        "G4U_DIFF_CREATED_START", "2026-01-01T00:00:00.000Z"
    )
    created_end = args.created_end or env_get(
        "G4U_DIFF_CREATED_END", "2026-06-30T23:59:59.999Z"
    )

    amap = load_action_map(MAP_PATH)
    stage_map: dict[str, str] = amap.get("stageTitleToActionTemplateId") or {}
    tag_map: dict[str, str] = amap.get("tagFlowTitleToActionTemplateId") or {}
    norm_stage_map: dict[str, tuple[str, str]] = {}
    for k, v in stage_map.items():
        nk = norm_key(k)
        if nk not in norm_stage_map:
            norm_stage_map[nk] = (k, v)

    picked = pick_xlsx_files(args.export_dir)
    if not picked:
        print(f"Nenhum xlsx reconhecido em {args.export_dir}", file=sys.stderr)
        return 1

    export_rows: list[dict[str, Any]] = []
    for kind, pth in picked:
        if kind == "Jurídico":
            export_rows.extend(read_jur_sheet(pth, tag_map))
        else:
            export_rows.extend(read_stage_sheet(kind, pth, stage_map, norm_stage_map))

    g4u_index: dict[str, dict[str, Any]] = {}
    if not args.dry_parse:
        base = env_get("G4U_API_BASE", "http://localhost:3001")
        client_id = env_get("CLIENT_ID") or env_get("client_id") or "revisaprev"
        token = env_get("G4U_ACCESS_TOKEN")
        if not token:
            email = env_get("G4U_ADMIN_EMAIL")
            pw = env_get("G4U_ADMIN_PASSWORD")
            if not email or not pw:
                print(
                    "Defina G4U_ACCESS_TOKEN ou G4U_ADMIN_EMAIL + G4U_ADMIN_PASSWORD",
                    file=sys.stderr,
                )
                return 1
            try:
                token = game4u_login(base, client_id, email, pw)
            except (URLError, HTTPError, RuntimeError) as e:
                print(f"Login falhou: {e}", file=sys.stderr)
                return 1
        statuses = ["PENDING", "DONE", "CANCELLED"]
        limit = int(env_get("G4U_USER_ACTION_SEARCH_LIMIT") or "500")
        max_pages = int(env_get("G4U_USER_ACTION_SEARCH_MAX_PAGES") or "80")
        udeals = sorted({str(er["deal_id"]) for er in export_rows if er.get("deal_id")})
        per_deal_env = env_get("G4U_DIFF_PER_DEAL", "").lower() in ("1", "true", "yes")
        if per_deal_env:
            print(
                f"G4U_DIFF_PER_DEAL: buscando user_action por delivery_id ({len(udeals)} deals únicos)…",
                file=sys.stderr,
            )
            raw = fetch_user_actions_per_deal(
                base,
                client_id,
                token,
                udeals,
                created_start,
                created_end,
                statuses,
                limit,
                max_pages,
            )
            g4u_index = index_by_integration_id(raw)
        else:
            try:
                raw = fetch_all_user_actions(
                    base, client_id, token, created_start, created_end, statuses, limit, max_pages
                )
            except RuntimeError as e:
                err_txt = str(e)
                if "delivery" in err_txt.lower() or "delivery_id" in err_txt.lower():
                    print(
                        "Busca global falhou; tentando por deal_id do export "
                        f"({len(udeals)} deals). Ou defina G4U_DIFF_PER_DEAL=1.",
                        file=sys.stderr,
                    )
                    raw = fetch_user_actions_per_deal(
                        base,
                        client_id,
                        token,
                        udeals,
                        created_start,
                        created_end,
                        statuses,
                        limit,
                        max_pages,
                    )
                    g4u_index = index_by_integration_id(raw)
                else:
                    print(err_txt, file=sys.stderr)
                    return 1
            else:
                g4u_index = index_by_integration_id(raw)

    ts = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
    out_path = args.out or (args.export_dir / f"diff-zoho-game4u-{ts}.csv")

    fieldnames = [
        "source_kind",
        "source_file",
        "sheet_row",
        "export_record_id",
        "deal_id",
        "deal_name",
        "stage_or_subject",
        "zoho_tag",
        "zoho_task_status",
        "modified_time",
        "fluxo",
        "expected_action_template_id",
        "expected_integration_id",
        "map_key_or_note",
        "found_in_game4u",
        "game4u_status",
        "game4u_id",
        "game4u_integration_id",
        "search_window",
    ]

    with out_path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        w.writeheader()
        for er in export_rows:
            exp_iid = er.get("expected_integration_id")
            hit = g4u_index.get(exp_iid) if exp_iid else None
            map_col = er.get("map_key", "") or er.get("map_note", "")
            w.writerow(
                {
                    "source_kind": er["source_kind"],
                    "source_file": er["source_file"],
                    "sheet_row": er["row"],
                    "export_record_id": er.get("export_record_id"),
                    "deal_id": er.get("deal_id"),
                    "deal_name": er.get("deal_name"),
                    "stage_or_subject": er.get("stage"),
                    "zoho_tag": er.get("zoho_tag", ""),
                    "zoho_task_status": er.get("zoho_task_status", ""),
                    "modified_time": er.get("modified"),
                    "fluxo": er.get("fluxo", ""),
                    "expected_action_template_id": er.get("expected_template"),
                    "expected_integration_id": exp_iid,
                    "map_key_or_note": map_col,
                    "found_in_game4u": "yes" if hit else "no",
                    "game4u_status": (hit.get("status") if hit else ""),
                    "game4u_id": (hit.get("id") if hit else ""),
                    "game4u_integration_id": (hit.get("integration_id") if hit else ""),
                    "search_window": f"{created_start} .. {created_end}",
                }
            )

    # Resumo no stderr + .md
    n = len(export_rows)
    found = sum(
        1
        for er in export_rows
        if er.get("expected_integration_id")
        and g4u_index.get(er["expected_integration_id"])
    )
    missing_tpl = sum(1 for er in export_rows if not er.get("expected_integration_id"))
    with_tpl = n - missing_tpl
    found_no = with_tpl - found
    print(
        f"Linhas export: {n} | com integration_id esperado e encontrado no Game4U: {found} | sem template/integration_id: {missing_tpl}",
        file=sys.stderr,
    )
    print(f"CSV: {out_path}", file=sys.stderr)

    md_path = out_path.with_suffix(".md")
    unmapped = Counter()
    for er in export_rows:
        note = (er.get("map_note") or "").strip()
        if note.startswith("UNMAPPED_STAGE:"):
            unmapped[note.split(":", 1)[1][:80]] += 1
        elif note.startswith("UNMAPPED_JUR"):
            unmapped[note[:100]] += 1
    lines = [
        "# Diff Zoho export × Game4U user_action",
        "",
        f"- Gerado em (UTC): `{datetime.now(timezone.utc).isoformat()}`",
        f"- Janela search: `{created_start}` … `{created_end}`",
        f"- Linhas no export: **{n}**",
        f"- Com `integration_id` esperado: **{with_tpl}**",
        f"- Encontrados no Game4U (`found_in_game4u=yes`): **{found}**",
        f"- Não encontrados (mas com id esperado): **{found_no}**",
        f"- Sem template / sem `integration_id`: **{missing_tpl}**",
        "",
        "Arquivo detalhado: `" + out_path.name + "`",
        "",
    ]
    if unmapped:
        lines.append("## Estágios / jur sem mapeamento (top 30)")
        lines.append("")
        for label, cnt in unmapped.most_common(30):
            lines.append(f"- **{cnt}** × `{label}`")
        lines.append("")
    md_path.write_text("\n".join(lines), encoding="utf-8")
    print(f"Resumo: {md_path}", file=sys.stderr)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
